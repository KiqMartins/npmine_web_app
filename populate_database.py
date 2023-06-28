import psycopg2
from openpyxl import load_workbook
from tqdm import tqdm

# Establish a connection to the database
conn = psycopg2.connect(
    host="localhost",
    port=5434,
    database="npmine",
    user="postgres",
    password="6243931"
)

# Create a cursor object to execute SQL queries
cursor = conn.cursor()

# Populate the Compounds table from XLSX
workbook_compounds = load_workbook('Table S1_total_unique_inchikeys.xlsx')
sheet_compounds = workbook_compounds.active

# Dictionary to store article_url to doi_id mapping
article_url_to_doi_id = {}

total_compounds = sheet_compounds.max_row - 1  # Exclude the header row

with tqdm(total=total_compounds, desc='Populating Compounds') as pbar:
    for row in sheet_compounds.iter_rows(min_row=2, values_only=True):
        if len(row) < 8:
            continue

        journal, smiles, article_url, inchi_key, exact_molecular_weight, pubchem_id, inchi, source = row[:8]

        # Check if the article_url is already mapped to a doi_id
        if article_url in article_url_to_doi_id:
            doi_id = article_url_to_doi_id[article_url]
        else:
            # Insert the DOI into the DOI table
            cursor.execute("INSERT INTO doi (doi) VALUES (%s) RETURNING id", (article_url,))
            doi_id = cursor.fetchone()[0]
            article_url_to_doi_id[article_url] = doi_id

        # Handle empty values
        exact_molecular_weight = float(exact_molecular_weight) if exact_molecular_weight else None
        pubchem_id = int(pubchem_id) if pubchem_id else None

        # Insert the compound into the Compounds table
        cursor.execute(
            "INSERT INTO compounds (journal, smiles, article_url, inchi_key, exact_molecular_weight, pubchem_id, inchi, source, user_id) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (journal, smiles, article_url, inchi_key, exact_molecular_weight, pubchem_id, inchi, source, None)
        )

        # Get the compound ID
        cursor.execute("SELECT id FROM compounds WHERE article_url = %s", (article_url,))
        compound_id = cursor.fetchone()[0]

        # Insert the compound ID and DOI ID into the doicomp table
        cursor.execute("INSERT INTO doicomp (doi_id, compounds_id) VALUES (%s, %s)", (doi_id, compound_id))

        # Update the progress bar
        pbar.update()

# Populate the Taxa table from XLSX
workbook_taxons = load_workbook('df_taxons_final.xlsx')
sheet_taxons = workbook_taxons.active

total_taxons = sheet_taxons.max_row - 1  # Exclude the header row

with tqdm(total=total_taxons, desc='Populating Taxa') as pbar:
    for row in sheet_taxons.iter_rows(min_row=2, values_only=True):
        if len(row) < 8:
            continue

        article_url, verbatim, odds, datasourceid, taxonid, classificationpath, classificationrank, matchtype = row[:8]

        # Check if the article_url is already mapped to a doi_id
        if article_url in article_url_to_doi_id:
            doi_id = article_url_to_doi_id[article_url]
        else:
            # Insert the DOI into the DOI table
            cursor.execute("INSERT INTO doi (doi) VALUES (%s) RETURNING id", (article_url,))
            doi_id = cursor.fetchone()[0]
            article_url_to_doi_id[article_url] = doi_id

        # Insert the taxon into the Taxa table
        cursor.execute(
            "INSERT INTO taxa (article_url, verbatim, odds, datasourceid, taxonid, classificationpath, classificationrank, matchtype, user_id) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (article_url, verbatim, odds, datasourceid, taxonid, classificationpath, classificationrank, matchtype, None)
        )

        # Get the taxon ID
        cursor.execute("SELECT id FROM taxa WHERE article_url = %s", (article_url,))
        taxon_id = cursor.fetchone()[0]

        # Insert the taxon ID and DOI ID into the doitaxa table
        cursor.execute("INSERT INTO doitaxa (doi_id, taxa_id) VALUES (%s, %s)", (doi_id, taxon_id))

        # Update the progress bar
        pbar.update()

# Commit the changes to the database
conn.commit()

# Close the cursor and connection
cursor.close()
conn.close()













